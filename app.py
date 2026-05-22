from __future__ import annotations

import argparse
import io
import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


PAGE_W, PAGE_H = landscape(A4)
FONT = "OrderFormNotoSansTC"
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(15 * 1024 * 1024)))
TEXT_REPLACEMENTS = {
    # Windows PMingLiU contains U+9039, but bundled Noto Sans TC does not.
    "逹": "達",
}
DETAIL_TOP = 410.56
DETAIL_BOTTOM = 58.0
DETAIL_HEADER_H = 19.52
DETAIL_LINE_H = 10.2
DETAIL_ROW_PAD = 4.0
DETAIL_MIN_ROW_H = 24.0

app = FastAPI(title="Order Helper", version="0.2.0")


@dataclass
class OrderRow:
    order_no: str
    item_no: str
    code: str
    name: str
    unit: str
    quantity: str
    vendor: str
    tel: str
    fax: str


@dataclass
class VendorPage:
    vendor: str
    tel: str
    fax: str
    page_no: int
    total_pages: int
    orders: list[OrderRow]


TERMS = {
    "order_no": "訂購單號",
    "item_no": "項次",
    "code": "料號",
    "name": "品名規格",
    "unit": "單位",
    "quantity": "訂購量",
    "vendor": "廠商",
    "tel": "電話",
    "fax": "傳真",
}


_APP_DIR = Path(__file__).resolve().parent

FONT_CANDIDATES = [
    os.getenv("ORDERHELPER_FONT_PATH", ""),
    # Bundled in the deployment zip. Bold first so all text renders bold.
    str(_APP_DIR / "fonts" / "NotoSansTC-Bold.ttf"),
    str(_APP_DIR / "fonts" / "NotoSansTC-VariableFont_wght.ttf"),
    str(_APP_DIR / "fonts" / "NotoSansTC-Regular.otf"),
    str(_APP_DIR / "fonts" / "NotoSansTC-Regular.ttf"),
    # Windows local dev.
    r"C:\Windows\Fonts\NotoSansTC-VF.ttf",
    r"C:\Windows\Fonts\NotoSansCJK-Regular.ttc",
    # Linux apt install fonts-noto-cjk (e.g. local docker).
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansTC-Regular.otf",
    "/usr/share/fonts/truetype/noto/NotoSansTC-Regular.ttf",
]


INDEX_HTML = """<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>義大醫院訂購單 PDF 產生器</title>
  <style>
    body { font-family: "Noto Sans TC", "Microsoft JhengHei", Arial, sans-serif; margin: 40px; color: #1f2933; }
    main { max-width: 720px; }
    label { display: block; margin: 18px 0 6px; font-weight: 700; }
    input[type=file], input[type=date] { width: 100%; box-sizing: border-box; padding: 10px; border: 1px solid #9aa5b1; border-radius: 4px; }
    button { margin-top: 22px; padding: 10px 18px; border: 0; border-radius: 4px; background: #205493; color: white; font-weight: 700; cursor: pointer; }
    p { line-height: 1.7; }
  </style>
</head>
<body>
  <main>
    <h1>義大醫院訂購單 PDF 產生器</h1>
    <p>上傳 Excel 後，系統會依廠商合併訂購明細；同一廠商多筆藥品會排在同一張訂單，並輸出單一 PDF。</p>
    <form method="post" action="/generate" enctype="multipart/form-data">
      <label for="excel">Excel 檔案</label>
      <input id="excel" name="excel" type="file" accept=".xlsx" required>
      <label for="order_date">訂貨日期</label>
      <input id="order_date" name="order_date" type="date">
      <button type="submit">產生 PDF</button>
    </form>
  </main>
</body>
</html>
"""


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    for source, target in TEXT_REPLACEMENTS.items():
        text = text.replace(source, target)
    return text


def _font_path() -> str:
    for candidate in FONT_CANDIDATES:
        if candidate and Path(candidate).exists():
            return candidate
    raise RuntimeError(
        "找不到可用的 Noto Sans TC/CJK 字型。"
        "請安裝 fonts-noto-cjk，或設定 ORDERHELPER_FONT_PATH 指向字型檔。"
    )


def _instance_variable_font(path: str, weight: int = 700) -> str:
    """If `path` is a variable font with a weight axis, return a cached static
    instance at the requested weight (default 700 = Bold). Otherwise return
    `path` unchanged. Silently falls back if fontTools is unavailable."""
    try:
        from fontTools.ttLib import TTFont as _FTFont
        from fontTools.varLib.instancer import instantiateVariableFont
    except ImportError:
        return path
    try:
        ft = _FTFont(path)
    except Exception:
        return path
    if "fvar" not in ft:
        return path  # not a variable font

    cache_dir = Path(os.getenv("TMPDIR", "/tmp"))
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{Path(path).stem}-static-w{weight}.ttf"
    if cache_path.exists():
        return str(cache_path)
    static = instantiateVariableFont(ft, {"wght": weight})
    static.save(str(cache_path))
    return str(cache_path)


def register_fonts() -> None:
    if FONT in set(pdfmetrics.getRegisteredFontNames()):
        return
    path = _font_path()
    weight = int(os.getenv("ORDERHELPER_FONT_WEIGHT", "700"))
    path = _instance_variable_font(path, weight=weight)
    suffix = Path(path).suffix.lower()
    if suffix == ".ttc":
        pdfmetrics.registerFont(TTFont(FONT, path, subfontIndex=0))
    else:
        pdfmetrics.registerFont(TTFont(FONT, path))


def _draw_string(c: canvas.Canvas, x: float, y: float, text: str) -> None:
    c.drawString(x, y, _text(text))


def _draw_right(c: canvas.Canvas, x: float, y: float, text: str) -> None:
    c.drawRightString(x, y, _text(text))


def _draw_center(c: canvas.Canvas, x: float, y: float, text: str) -> None:
    c.drawCentredString(x, y, _text(text))


def _draw_center_bold(c: canvas.Canvas, x: float, y: float, text: str) -> None:
    _draw_center(c, x, y, text)
    _draw_center(c, x + 0.35, y, text)


def _draw_fit_string(
    c: canvas.Canvas,
    x: float,
    y: float,
    text: str,
    max_width: float,
    font_size: float,
    min_font_size: float = 7.0,
) -> None:
    value = _text(text)
    size = font_size
    while size > min_font_size and pdfmetrics.stringWidth(value, FONT, size) > max_width:
        size -= 0.5
    c.setFont(FONT, size)
    _draw_string(c, x, y, value)


def _vendor_key(order: OrderRow) -> str:
    return order.vendor


def _find_order_sheet(workbook):
    for ws in workbook.worksheets:
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            values = [_text(v) for v in row]
            if any(TERMS["order_no"] in v for v in values) and any(TERMS["vendor"] in v for v in values):
                index = {v: i for i, v in enumerate(values) if v}
                cols = {}
                for key, term in TERMS.items():
                    match = next((h for h in index if term in h), None)
                    if not match:
                        raise ValueError(f"Excel 缺少欄位：{term}")
                    cols[key] = index[match]
                return ws, row_no, cols
    raise ValueError("找不到包含「訂購單號」與「廠商」欄位的訂單工作表。")


def read_orders(source) -> list[OrderRow]:
    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        ws, header_row, cols = _find_order_sheet(wb)
        orders: list[OrderRow] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            order_no = _text(row[cols["order_no"]])
            if not order_no:
                continue
            orders.append(
                OrderRow(
                    order_no=order_no,
                    item_no=_text(row[cols["item_no"]]),
                    code=_text(row[cols["code"]]),
                    name=_text(row[cols["name"]]),
                    unit=_text(row[cols["unit"]]),
                    quantity=_text(row[cols["quantity"]]),
                    vendor=_text(row[cols["vendor"]]),
                    tel=_text(row[cols["tel"]]),
                    fax=_text(row[cols["fax"]]),
                )
            )
        return orders
    finally:
        wb.close()


def infer_order_date(filename: str, orders: list[OrderRow]) -> str:
    candidates = [filename, *(o.order_no for o in orders[:5])]
    for text in candidates:
        match = re.search(r"(\d{3})(\d{2})(\d{2})", text)
        if match:
            roc_year, month, day = map(int, match.groups())
            return f"{roc_year + 1911:04d}-{month:02d}-{day:02d}"
    return date.today().isoformat()


def _fit_text(text: str, max_width: float, font_name: str, font_size: float) -> list[str]:
    lines: list[str] = []
    current = ""
    for char in text:
        trial = current + char
        if current and pdfmetrics.stringWidth(trial, font_name, font_size) > max_width:
            lines.append(current)
            current = char
        else:
            current = trial
    if current:
        lines.append(current)
    return lines[:5]


def _fit_detail_name(text: str) -> list[str]:
    return _fit_text(text, 345, FONT, 9) or [""]


def _detail_row_height(order: OrderRow) -> float:
    line_count = max(1, len(_fit_detail_name(order.name)))
    return max(DETAIL_MIN_ROW_H, DETAIL_ROW_PAD * 2 + line_count * DETAIL_LINE_H)


def _paginate_vendor_orders(orders: list[OrderRow]) -> list[list[OrderRow]]:
    pages: list[list[OrderRow]] = []
    current: list[OrderRow] = []
    used_height = 0.0
    capacity = DETAIL_TOP - DETAIL_HEADER_H - DETAIL_BOTTOM
    for order in orders:
        row_height = _detail_row_height(order)
        if current and used_height + row_height > capacity:
            pages.append(current)
            current = []
            used_height = 0.0
        current.append(order)
        used_height += row_height
    if current:
        pages.append(current)
    return pages


def build_vendor_pages(orders: list[OrderRow]) -> list[VendorPage]:
    grouped: dict[str, list[OrderRow]] = {}
    keys: list[str] = []
    for order in orders:
        key = _vendor_key(order)
        if key not in grouped:
            grouped[key] = []
            keys.append(key)
        grouped[key].append(order)

    pages: list[VendorPage] = []
    for key in keys:
        vendor_orders = grouped[key]
        first = vendor_orders[0]
        chunks = _paginate_vendor_orders(vendor_orders)
        for page_no, chunk in enumerate(chunks, 1):
            pages.append(
                VendorPage(
                    vendor=key,
                    tel=first.tel,
                    fax=first.fax,
                    page_no=page_no,
                    total_pages=len(chunks),
                    orders=chunk,
                )
            )
    return pages


def _draw_rects(c: canvas.Canvas, orders: list[OrderRow]) -> list[float]:
    c.setLineWidth(0.4)
    row_heights = [_detail_row_height(order) for order in orders]
    detail_body_h = max(45.44, sum(row_heights))
    detail_bottom = DETAIL_TOP - DETAIL_HEADER_H - detail_body_h
    rects = [
        (21.76, 504.24, 801.0, 24.0),
        (21.76, 430.04, 801.0, 74.23),
        (21.76, 410.56, 801.0, 19.52),
        (63.0, 410.52, 105.0, 19.52),
        (240.76, 410.52, 75.76, 19.52),
        (678.76, 410.52, 75.76, 19.52),
        (270.0, 430.0, 293.24, 74.23),
        (21.76, detail_bottom, 801.0, detail_body_h),
    ]
    for rect in rects:
        c.rect(*rect, stroke=1, fill=0)
    y = DETAIL_TOP - DETAIL_HEADER_H
    for row_height in row_heights[:-1]:
        y -= row_height
        c.line(21.76, y, 822.76, y)
    for x, y1, y2 in [
        (63.04, DETAIL_TOP, detail_bottom),
        (168.04, DETAIL_TOP, detail_bottom),
        (240.68, DETAIL_TOP, detail_bottom),
        (316.48, DETAIL_TOP, detail_bottom),
        (678.76, DETAIL_TOP, detail_bottom),
        (754.48, DETAIL_TOP, detail_bottom),
    ]:
        c.line(x, y1, x, y2)
    return row_heights


def _draw_static(c: canvas.Canvas, page_no: int, total_pages: int) -> None:
    c.setFont(FONT, 20)
    _draw_center_bold(c, 421, 566.4, "義大醫療財團法人義大醫院")
    c.setFont(FONT, 18)
    _draw_center_bold(c, 421, 540.1, "藥品訂購單")
    c.setFont(FONT, 14)
    _draw_string(c, 757.6, 538.2, str(page_no))
    _draw_string(c, 778.0, 536.7, "/")
    _draw_string(c, 800.8, 538.2, str(total_pages))
    c.setFont(FONT, 10)
    _draw_string(c, 28.8, 535.8, "報表代碼：INV_APP_07")

    c.setFont(FONT, 12)
    _draw_string(c, 30.8, 512.7, "廠商名稱")
    _draw_string(c, 389.3, 512.7, "FAX：")
    c.rect(605.8, 511.2, 7.0, 7.0, stroke=1, fill=0)
    _draw_string(c, 616.0, 512.7, "mail訂貨日期")

    c.setFont(FONT, 10)
    _draw_string(c, 30.8, 488.6, "發票抬頭：義大醫療財團法人義大醫院")
    _draw_string(c, 30.8, 476.6, "發票地址：高雄市燕巢區角宿里義大路1號")
    _draw_string(c, 30.8, 464.6, "統一編號：25886456")
    _draw_string(c, 30.8, 452.6, "醫療機構代碼：1142120001")
    _draw_string(c, 30.8, 440.6, "**管證字號：QHP101000003")
    _draw_string(c, 279.7, 485.6, "交貨地址：高雄市燕巢區角宿里義大路1號")
    _draw_string(c, 279.7, 473.6, "□藥庫(B1F)   □_______")
    _draw_string(c, 279.7, 461.6, "聯絡電話：07-6150011#6226.6225.6224(藥庫)林藥師")
    _draw_string(c, 279.7, 449.6, "傳真：07-6154431")
    c.setFont(FONT, 9)
    _draw_string(c, 578.9, 490.3, "※備註：")
    _draw_string(c, 578.9, 479.3, "1.發票與出貨單請隨貨附上。")
    _draw_string(c, 578.9, 468.2, "2.請開立三聯式發票。")
    _draw_string(c, 578.9, 457.2, "3.請附8元回郵信封(郵寄折讓單)。")
    _draw_string(c, 578.9, 446.2, "4.首次交易或未申請匯款者請與採購課人員接洽辦理。")

    c.setFont(FONT, 12)
    _draw_string(c, 30.4, 417.1, "序號")
    _draw_string(c, 72.0, 417.1, "訂購編號")
    _draw_string(c, 181.5, 417.1, "料號項次")
    _draw_string(c, 256.5, 417.1, "料號")
    _draw_string(c, 432.0, 417.1, "品名規格")
    _draw_string(c, 691.0, 417.1, "計價單位")
    _draw_string(c, 766.0, 417.1, "訂購量")


def _draw_vendor_page(c: canvas.Canvas, page: VendorPage, order_date: str, row_heights: list[float]) -> None:
    c.setFont(FONT, 12)
    _draw_fit_string(c, 88.6, 513.4, page.vendor, 98.0, 12)
    c.setFont(FONT, 12)
    _draw_fit_string(c, 191.0, 512.7, f"TEL：{page.tel}", 190.0, 12)
    c.setFont(FONT, 12)
    _draw_fit_string(c, 423.0, 512.7, page.fax, 136.0, 12)
    c.setFont(FONT, 12)
    _draw_string(c, 747.6, 511.9, order_date)

    y_top = DETAIL_TOP - DETAIL_HEADER_H
    for idx, (order, row_height) in enumerate(zip(page.orders, row_heights), 1):
        y = y_top - 14.0
        c.setFont(FONT, 9)
        _draw_center(c, 42.4, y, str(idx))
        _draw_string(c, 68.0, y, order.order_no)
        _draw_string(c, 181.5, y, order.item_no)
        _draw_string(c, 259.5, y, order.code)
        _draw_string(c, 691.0, y, order.unit)
        _draw_right(c, 812.0, y, order.quantity)

        for line_idx, line in enumerate(_fit_detail_name(order.name)):
            _draw_string(c, 321.4, y - line_idx * DETAIL_LINE_H, line)
        y_top -= row_height


def build_pdf(orders: list[OrderRow], output, order_date: str) -> None:
    if not orders:
        raise ValueError("Excel 沒有可輸出的訂單資料。")
    register_fonts()
    pages = build_vendor_pages(orders)
    c = canvas.Canvas(output, pagesize=(PAGE_W, PAGE_H))
    for page in pages:
        row_heights = _draw_rects(c, page.orders)
        _draw_static(c, page.page_no, page.total_pages)
        _draw_vendor_page(c, page, order_date, row_heights)
        c.showPage()
    c.save()


def build_pdf_from_excel(excel_source, output, filename: str = "orders.xlsx", order_date: str | None = None) -> tuple[int, str, int]:
    orders = read_orders(excel_source)
    final_date = order_date or infer_order_date(filename, orders)
    build_pdf(orders, str(output) if isinstance(output, Path) else output, final_date)
    return len(orders), final_date, len({_vendor_key(order) for order in orders})


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return INDEX_HTML


@app.get("/health")
def health() -> JSONResponse:
    return JSONResponse({"status": "ok"})


@app.post("/generate")
async def generate_pdf(excel: UploadFile = File(...), order_date: str | None = Form(default=None)):
    filename = Path(excel.filename or "orders.xlsx").name
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 檔案。")

    content = await excel.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"檔案太大，限制為 {MAX_UPLOAD_BYTES // 1024 // 1024} MB。")

    try:
        output = io.BytesIO()
        count, final_date, vendor_count = build_pdf_from_excel(io.BytesIO(content), output, filename=filename, order_date=order_date or None)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    output.seek(0)
    safe_name = f"{Path(filename).stem}_訂購單_{final_date}_{vendor_count}廠商_{count}筆.pdf"
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_name)}"}
    return StreamingResponse(output, media_type="application/pdf", headers=headers)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("訂購單_output.pdf"))
    parser.add_argument("--date")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()
    if args.input:
        count, final_date, vendor_count = build_pdf_from_excel(args.input, args.output, filename=args.input.name, order_date=args.date)
        print(f"wrote {args.output} ({vendor_count} vendors, {count} rows, date {final_date})")
    else:
        import uvicorn

        uvicorn.run("app:app", host=args.host, port=args.port)


if __name__ == "__main__":
    main()
